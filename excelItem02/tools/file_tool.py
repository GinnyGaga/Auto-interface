from config import base_path, cell_config
import os
import openpyxl
import json


class FileTool:
    # 1. 初始化
    def __init__(self, filename):
        # 1. 动态文件路径
        self.filename = base_path + os.sep + "data" + os.sep + filename
        print("要的打开的文件为：", self.filename)
        # 2. 打开文件 获取workbook对象  load_workbook("excel文件名")
        self.workbook = openpyxl.load_workbook(self.filename)
        # 3. 获取sheet表单对象，获取用例所在的表单
        self.sheet = self.workbook[self.workbook.sheetnames[0]]
        # 4. 获取总行数max_row (读取数据，遍历数据使用),总列数为max_column
        self.row = self.sheet.max_row
        print("总行数数据为：", self.row)
        # 5. 记录对照组结果
        self.case = {}

    # 2. 读取Excel
    def read_excel(self):
        # 1. 新建空列表 （存储每行数据）
        case = list()
        # 2. 遍历每行数据，左闭右开
        for i in range(2, self.row + 1):
            # 新建空字典 （存储每行数据）
            data = dict()
            # 判断是否执行,若单元格的值是Run，则执行
            if self.sheet.cell(i, cell_config.get("is_run")).value == "Run":
                try:
                    # 读取数据 追加到字典 sheet.cell(x,y).value读取sheet对象行x列y单元格的值
                    data['path'] = self.sheet.cell(i, cell_config["path"]).value
                    data['method'] = str(self.sheet.cell(i, cell_config["method"]).value).lower()
                    data['headers'] = eval(self.sheet.cell(i, cell_config["headers"]).value)
                    data['param_type'] = self.sheet.cell(i, cell_config["param_type"]).value
                    data['params'] = self.sheet.cell(i, cell_config["params"]).value
                    data['expect'] = self.sheet.cell(i, cell_config["expect"]).value
                    data['expect'] = json.loads(data['expect'])  # 字符串转换为json对象
                    # 记录每行数据执行结果的行与列，执行完写入测试结果（后面使用）；
                    data['x_y'] = [i, cell_config.get("result")]
                    # 将字典追加到列表中
                    case.append(data)
                    # 将读取结果写入excel中，表示该框架能正常读取excel文件
                    self.write_excel([i, cell_config.get("desc")], "数据读取完成～！")
                    self.case = case
                    print("读取的数据为：", self.case)

                except Exception as e:
                    self.write_excel([i, cell_config.get("desc")], str(e))
                    self.workbook.close()

        # 3. 关闭excel文件句柄
        self.workbook.close()

    # 3. 写入Excel
    def write_excel(self, x_y, msg):
        try:
            # x_y参数的格式为列表 如: [2,5]
            self.sheet.cell(x_y[0], x_y[1]).value = msg
        except Exception as e:
            # x_y参数的格式为列表 如: [2,5]
            self.sheet.cell(x_y[0], x_y[1]).value = e
        finally:
            # 保存excel
            self.workbook.save(self.filename)

    # 4. 读取Json
    def read_json(self, file_name="case.json"):
        file_name = base_path + os.sep + "data" + os.sep + file_name
        with open(file_name, "r", encoding="utf-8") as f:
            return json.load(f)

    # 5. 写入Json文件
    def write_json(self, file_name):
        file_name = base_path + os.sep + "data" + os.sep + file_name
        with open(file_name, "w", encoding="utf-8")as f:
            json.dump(self.case, f, indent=4, ensure_ascii=False)  # 把json object转换成json str，写入文件case.json


if __name__ == '__main__':
    tool = FileTool("Case01.xlsx")
    tool.read_excel()
    tool.write_json("case.json")  # 读写分离执行
