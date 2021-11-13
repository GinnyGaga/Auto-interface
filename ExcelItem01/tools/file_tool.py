
import os
import openpyxl

from config import base_path, cell_config
import json


class FileTool:
    # 1、初始化
    def __init__(self, filename):
        # 1、动态文件路径  os.sep根据你所处的平台，自动采用相应的分隔符号
        self.filename = base_path + os.sep + 'data' + os.sep + filename
        print('要打开的文件为：', self.filename)
        # 2、打开文件，获取workbook对象
        self.workbook = openpyxl.load_workbook(self.filename)
        # 3、获取sheet表单
        self.sheet = self.workbook[self.workbook.sheetnames[0]]
        # 4、获取总行数(读取数据，遍历数据使用)
        self.row = self.sheet.max_row
        print('总行数数据为：', self.row)

    # 2、读取Excel
    def read_excel(self):
        # 1、新建空列表 （存储每行数据）
        case = list()
        # 2、遍历每行数据
        for i in range(2, self.row + 1):
            # 新建空字典（存储每行数据）
            data = dict()
            # 判断是否执行
            # if self.sheet.cell(i, cell_config.get("result")).value == "是":
            if self.sheet.cell(i, cell_config.get("is_run")).value == "是":
                # 读取数据 追加到字典
                try:
                    data['path'] = self.sheet.cell(i, cell_config.get('path')).value
                    data['method'] = str(self.sheet.cell(i, cell_config.get('method')).value).lower()
                    data['headers'] = self.sheet.cell(i, cell_config.get('headers')).value
                    data['param_type'] = self.sheet.cell(i, cell_config.get('param_type')).value
                    data['params'] = self.sheet.cell(i, cell_config.get('params')).value
                    data['expect'] = self.sheet.cell(i, cell_config.get('expect')).value
                    # 记录每行数据的 行和列，执行完写入结果使用
                    data['x_y'] = [i, cell_config.get("is_run")]  # 读取表格中"is_run"的位置
                    # 将字典追加到列表
                    case.append(data)
                    # 将读取结果写入excel中
                    self.write_excel([i, cell_config.get("desc")], "数据读取完成")
                except Exception as e:
                    self.write_excel([i, cell_config.get("desc")], e)
        # 3.将列表数据写入json
        self.write_json(case, "case.json")
        print("读取列表数据为：", case)

    # 3、写入Excel
    def write_excel(self, x_y, msg):
        try:
            # x_y参数的格式为列表 如：[2,5]
            self.sheet.cell(x_y[0], x_y[1]).value = msg
        except Exception as e:
            # x_y参数的格式为列表 如：[2,5]
            self.sheet.cell(x_y[0], x_y[1]).value = e
        finally:
            # 保存excel
            self.workbook.save(self.filename)

    # 4、读取Json
    def read_json(self, file_name='case.json'):
        file_name = base_path + os.sep + "data" + os.sep + file_name
        with open(file_name, "r", encoding="utf-8") as f:
            return json.load(f)

    # 5、写入Json
    def write_json(self,case,file_name):
        file_name = base_path + os.sep + "data" + os.sep + file_name
        with open(file_name,"w",encoding="utf-8") as f:
            json.dump(case,f,indent=4,ensure_ascii=False)  # ensure_ascii为假，则不转义中文字符


if __name__ == '__main__':
    d = FileTool('Case01.xlsx')
    d.read_excel()
    # d.write_excel([3, 6], "POST")  # 把表格里的第3行6列的值改成POST
    print(d.read_json()) # 已默认传入参数
